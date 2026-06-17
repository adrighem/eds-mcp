import asyncio
import logging
from pathlib import Path
from typing import Callable

logger = logging.getLogger(__name__)

MAX_FILE_SIZE_MB = 5
MAX_CONTENT_CHARS = 30000
TEXT_EXTENSIONS = {".txt", ".md", ".csv", ".json"}
Parser = Callable[[str], str]


def parse_pdf(file_path: str) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        page_text = [(page.extract_text() or "") for page in reader.pages]
        return "\n".join(page_text).strip()
    except ImportError:
        return "Error: pypdf library not installed."
    except Exception as e:
        return f"Error parsing PDF: {e}"


def parse_docx(file_path: str) -> str:
    try:
        from docx import Document
        doc = Document(file_path)
        return "\n".join(para.text for para in doc.paragraphs).strip()
    except ImportError:
        return "Error: python-docx library not installed."
    except Exception as e:
        return f"Error parsing DOCX: {e}"


def parse_xlsx(file_path: str) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        lines = []
        for sheet in wb.worksheets:
            lines.append(f"### Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                lines.append(" | ".join(str(cell) if cell is not None else "" for cell in row))
            lines.append("")
        return "\n".join(lines).strip()
    except ImportError:
        return "Error: openpyxl library not installed."
    except Exception as e:
        return f"Error parsing XLSX: {e}"


def parse_text(file_path: str) -> str:
    try:
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            return f.read()
    except Exception as e:
        return f"Error reading text file: {e}"


DOCUMENT_PARSERS: dict[str, Parser] = {
    ".pdf": parse_pdf,
    ".docx": parse_docx,
    ".xlsx": parse_xlsx,
    **{ext: parse_text for ext in TEXT_EXTENSIONS},
}


def supported_document_formats() -> str:
    """Return supported extensions in a stable human-readable order."""
    return ", ".join(sorted(DOCUMENT_PARSERS))


def truncate_for_context(content: str, max_chars: int = MAX_CONTENT_CHARS) -> str:
    """Keep parsed document content within a predictable context budget."""
    if len(content) <= max_chars:
        return content
    return content[:max_chars] + "\n\n[... content truncated due to length ...]"


async def parse_document_logic(file_path: str) -> str:
    path = Path(file_path)
    if not path.exists():
        return f"Error: File not found at {file_path}"
    
    # Check file size (limit parsing to 5MB to avoid context issues)
    size_mb = path.stat().st_size / (1024 * 1024)
    if size_mb > MAX_FILE_SIZE_MB:
        return f"Error: File is too large ({size_mb:.1f} MB). Max size for parsing is {MAX_FILE_SIZE_MB} MB."

    ext = path.suffix.lower()
    parser = DOCUMENT_PARSERS.get(ext)
    if not parser:
        return f"Error: Unsupported document format '{ext}'. Supported: {supported_document_formats()}"
    
    content = await asyncio.to_thread(parser, str(path))
    return truncate_for_context(content, MAX_CONTENT_CHARS)
